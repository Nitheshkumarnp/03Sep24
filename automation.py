#   This file is working well and used for all the 4 tabs it's original code taken from adding-changes-on-working-fine-code.py
#   which present in 22Aug24 folder of Ganesh sir system code.

#   In this file, I am removing unwanted codes and updating the comments.

import logging
import os
import sys
import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl

# Configuring logs
logging.basicConfig(
    filename='logs.log',  # Log file location
    format='%(asctime)s - %(levelname)s - %(message)s',
    level=logging.INFO  # Set the logging level
)

# Read data from excel file and store it in list
def read_ids_from_excel(file_path, lower_limit, upper_limit):
    '''
    This method will read data from excel file. Based on the lower and upper limit, data from excel will be added to ids list and ids will be returned as output of the method.

    Parameters:

        file_path = path of the selected file,
        lower_limit = minimum row of the excel file,
        upper_limit = maximum row of the excel file
    
    '''
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    ids = []
    for row in sheet.iter_rows(min_row=lower_limit, max_row=upper_limit, max_col=1, values_only=True):  # Assuming IDs are in the first column
        ids.append(row[0])
    return ids

# This is used to switch the mode from debug and development automatically based on system arguments.
debug_mode = True if len(sys.argv) < 2 else False

# required inputs
if debug_mode:
    # Hardcoding data required to test manually without starting the application
    excel_file_path = "BSP_new_excel_data.xlsx"
    lower_limit = 2
    upper_limit = 60
    timeToLoad = 150
    fileName = "test_data.xlsx"
    tab = "bsp-travcom"
else:
    # Original data comes from api request
    excel_file_path = sys.argv[1]
    lower_limit = int(sys.argv[2])
    upper_limit = int(sys.argv[3])
    timeToLoad = int(sys.argv[4])
    fileName = sys.argv[5]
    tab = sys.argv[6]

# Gets list of id values come from method's output
ids_to_search = read_ids_from_excel(excel_file_path, lower_limit, upper_limit)

# Adding id list to logs and removing duplicates
logging.info(f'Total Number of ids = {len(ids_to_search)}')
logging.info(ids_to_search)
ids_to_search = list(dict.fromkeys(ids_to_search))
logging.info(f'Unique number of ids = {len(ids_to_search)}')
logging.info(ids_to_search)

# Initialize the WebDriver
# If the web driver can be accessed by any location, no need to give the path.
# chrome_driver_path = "C:\\Users\\ganesh.ss\\Desktop\\chromedriver-win64\\chromedriver.exe"  # Replace with the actual path
# driver = webdriver.Chrome(service=Service(chrome_driver_path))
driver = webdriver.Chrome()

# URL of the web page
url = "https://app.powerbi.com/home?ctid=b1e9c207-e901-43b7-a133-9d42b486216d&experience=power-bi"

# Open the web page
driver.get(url)

# Maximize the screen size
driver.maximize_window()

# First mail id
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="email"]')))
firstMailId = driver.find_element(By.XPATH, '//*[@id="email"]')
firstMailId.send_keys('ganesh.ss@in.fcm.travel')
firstMailId.send_keys(Keys.RETURN)

# Second mail id
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="okta-signin-username"]')))
secondMailId = driver.find_element(By.XPATH, '//*[@id="okta-signin-username"]')
secondMailId.send_keys('ganesh.ss@fcmin.com')
secondMailId.send_keys(Keys.TAB)

# Second mail id password
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="okta-signin-password"]')))
secondMailId = driver.find_element(By.XPATH, '//*[@id="okta-signin-password"]')
secondMailId.send_keys('Ganu@87654321')
secondMailId.send_keys(Keys.RETURN)

# Push Button
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="form63"]/div[2]/input')))
pushButton = driver.find_element(By.XPATH, '//*[@id="form63"]/div[2]/input')
pushButton.click()

# Continue Button 1
WebDriverWait(driver, timeToLoad).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="idSIButton9"]')))
continueButton1 = driver.find_element(By.XPATH, '//*[@id="idSIButton9"]')
continueButton1.click()

# Continue Button 2
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="idSIButton9"]')))
continueButton2 = driver.find_element(By.XPATH, '//*[@id="idSIButton9"]')
continueButton2.click()

# Power BI Tabs
WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, '//*[@id="content"]/tri-shell/tri-item-renderer/tri-extension-page-outlet/div[2]/home/div/recommended-content-widget/trident-recommended-section/trident-recommended-card-list/div/trident-recommended-section-card[1]/tri-card/div/div/tri-svg-icon/img')))
powerBITabs = driver.find_element(By.XPATH, '//*[@id="content"]/tri-shell/tri-item-renderer/tri-extension-page-outlet/div[2]/home/div/recommended-content-widget/trident-recommended-section/trident-recommended-card-list/div/trident-recommended-section-card[1]/tri-card/div/div/tri-svg-icon/img')
powerBITabs.click()

# Common automation Script for all the 4 tabs
def automation(driver, tabNumber, inputFieldPath, tableHeaderPath, totalNoOfColumns, tableHeaderWithData):

    # This try block is implemented to avoid the time and data loss, while uncertain stop in automation function
    try:
        # This set of codes is to navigate to specific tab
        logging.info('Entering to automation method')
        WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.TAG_NAME, 'mat-action-list')))
        tabContainer = driver.find_element(By.TAG_NAME, 'mat-action-list')
        tabList = tabContainer.find_elements(By.TAG_NAME, 'button')
        tabList[tabNumber].click()
        time.sleep(5)

        # Iterating ids
        total_rows = 0
        for index, search_id in enumerate(ids_to_search):
            if search_id == None:
                break

            # Navigating to input field to enter the ids and switch to default content.
            WebDriverWait(driver, timeToLoad).until(EC.visibility_of_element_located((By.TAG_NAME, 'visual-container')))
            visualContainers = driver.find_elements(By.TAG_NAME, 'visual-container')
            iframe = visualContainers[inputFieldPath].find_element(By.TAG_NAME, 'iframe')
            driver.switch_to.frame(iframe)
            search_input = driver.find_element(By.TAG_NAME, 'input')
            search_input.clear()
            search_input.send_keys(search_id)
            search_input.send_keys(Keys.RETURN)
            driver.switch_to.default_content()
            time.sleep(2)

            # Converting table header content to list, To find number of columns
            WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.XPATH, tableHeaderPath)))
            tableHeaderBeforeScroll = driver.find_element(By.XPATH, tableHeaderPath)
            tableHeaderContent = tableHeaderBeforeScroll.text
            tableHeaderList = tableHeaderContent.split('\n')

            # removing the empty column header which comes at end unnecessarily
            while ' ' in tableHeaderList:
                tableHeaderList.remove(' ')

            # Calculating number of columns required before and after scroll
            tableColumnsBeforeScroll = len(tableHeaderList) - 1 if tableHeaderList[0] == 'Row Selection' else len(tableHeaderList)
            tableColumnsRequiredAfterScroll = totalNoOfColumns - tableColumnsBeforeScroll

            # Checking for scroll required or not
            if tableColumnsRequiredAfterScroll:
                # Scroll required flow

                # Setting and moving scroll bar to avoid stale element issue
                WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME, 'scroll-bar-part-bar')))
                scroll_bar = driver.find_element(By.CLASS_NAME, 'scroll-bar-part-bar')
                driver.execute_script("arguments[0].scrollIntoView();", scroll_bar)
                action = ActionChains(driver)
                action.click_and_hold(scroll_bar).move_by_offset(5, 0).release().perform()
                action.click_and_hold(scroll_bar).move_by_offset(-5, 0).release().perform()

                # Collecting data from tables before scroll
                tableDataBeforeScroll = []
                tableValuesBeforeScroll = []
                WebDriverWait(driver, 60).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, 'pivotTableCellWrap')))
                tableDataBeforeScroll = driver.find_elements(By.CLASS_NAME, 'pivotTableCellWrap')
                
                # Calculating number of rows, data present
                noOfRows = int(len(tableDataBeforeScroll) / tableColumnsBeforeScroll)

                total_rows += (noOfRows - 1)

                # Add logs for reference
                logging.info(f'Index = {index}, Search id = {search_id}, Rows = {noOfRows-1}, Total rows = {total_rows}')

                # Checking for empty values and avoid scrolling for empty tables
                if noOfRows > 1:

                    # Adding table values to list
                    temp = []
                    if len(tableDataBeforeScroll) > tableColumnsBeforeScroll:
                        for i in range(len(tableDataBeforeScroll)):
                            temp.append(tableDataBeforeScroll[i].text)
                    tableValuesBeforeScroll.extend(temp)

                    # Setting and moving scroll bar to navigate to right end as well as to avoid stale element issue
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME, 'scroll-bar-part-bar')))
                    scroll_bar = driver.find_element(By.CLASS_NAME, 'scroll-bar-part-bar')
                    driver.execute_script("arguments[0].scrollIntoView();", scroll_bar)
                    time.sleep(0.5)
                    action = ActionChains(driver)
                    action.click_and_hold(scroll_bar).move_by_offset(500, 0).release().perform()
                    action.click_and_hold(scroll_bar).move_by_offset(-5, 0).release().perform()

                    # Collecting data from table after scroll
                    tableDataAfterScroll = []
                    tableValuesAfterScroll = []
                    WebDriverWait(driver, 60).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, 'pivotTableCellWrap')))
                    tableDataAfterScroll = driver.find_elements(By.CLASS_NAME, 'pivotTableCellWrap')

                    # Calculating required table columns after scroll
                    tableColumnsAfterScroll = int(len(tableDataAfterScroll) / noOfRows)

                    # Adding table values to list
                    temp = []
                    if len(tableDataAfterScroll) > tableColumnsAfterScroll:
                        for i in range(len(tableDataAfterScroll)):
                            temp.append(tableDataAfterScroll[i].text)
                    tableValuesAfterScroll.extend(temp)

                    # Moving scroll bar to initial position
                    WebDriverWait(driver, 15).until(EC.visibility_of_element_located((By.CLASS_NAME, 'scroll-bar-part-bar')))
                    scroll_bar = driver.find_element(By.CLASS_NAME, 'scroll-bar-part-bar')
                    driver.execute_script("arguments[0].scrollIntoView();", scroll_bar)
                    action = ActionChains(driver)
                    action.click_and_hold(scroll_bar).move_by_offset(-500, 0).release().perform()

                    # Merging both the table values before and after values and converting to single list
                    first_list = [tableValuesBeforeScroll[i:i+tableColumnsBeforeScroll] for i in range(0, len(tableValuesBeforeScroll), tableColumnsBeforeScroll)]
                    second_list = [tableValuesAfterScroll[i:i+tableColumnsAfterScroll] for i in range(0, len(tableValuesAfterScroll), tableColumnsAfterScroll)]
                    merged_list = [sublist1 + sublist2[-tableColumnsRequiredAfterScroll:] for sublist1, sublist2 in zip(first_list, second_list)]
                    combined_list = [item for sublist in merged_list for item in sublist]
                    if len(combined_list) > totalNoOfColumns:
                        for table_value in combined_list[totalNoOfColumns:]:
                            tableHeaderWithData.append(table_value)
            else:
                # Scroll not required flow 

                # Getting table data available
                WebDriverWait(driver, 15).until(EC.visibility_of_all_elements_located((By.CLASS_NAME, 'pivotTableCellWrap')))
                tableData = driver.find_elements(By.CLASS_NAME, 'pivotTableCellWrap')

                # Calculating number of rows, data present
                noOfRows = int(len(tableData) / tableColumnsBeforeScroll)

                total_rows += (noOfRows - 1)

                # Adding table values to list
                temp = []
                if len(tableData) > tableColumnsBeforeScroll:
                    for table_value in tableData[totalNoOfColumns:]:
                        temp.append(table_value.text)
                    tableHeaderWithData.extend(temp)

        # Close the WebDriver
        driver.quit()

        # Logging and Printing overall result for better understanding
        logging.info(f'Id Number = {index}, Total Rows = {total_rows} are saved in excel file')
        print(f'Id Number = {index}, Total Rows = {total_rows} are saved in excel file')

        # Returning the all the datas in single list as tableHeaderWithData
        return tableHeaderWithData

    except:
        # Close the WebDriver
        driver.quit()

        # Logging and Printing overall result for better understanding
        logging.info(f'Id Number = {index}, Total Rows = {total_rows} are saved in excel file')
        print(f'Id Number = {index}, Total Rows = {total_rows} are saved in excel file')

        # Returning the all the datas in single list as tableHeaderWithData
        return tableHeaderWithData

# Setting total number of columns as global variable which is used in inserting of data in excel file.
totalNoOfColumns = 0

# find the tab and setting required inputs for automation method and calling it.
def findTab(tab):
    global totalNoOfColumns

    if tab == 'lcc-airlines':
        tabNumber = 1
        inputFieldPath = 2
        tableHeaderPath = '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[6]/transform/div/div[3]/div/div/visual-modern/div/div/div[2]/div[1]/div[1]/div/div/div'
        totalNoOfColumns = 11
        tableHeaderWithData = ["TYPE", "BRANCH", "Transation Date", "RecordLocator", "Airline Name", "TKTT Type", "Name1", "P Code", "Sum of AMOUNT", "MAIN PNR", "LOGIN ID"]
        
    elif tab == 'lcc-travcom':
        tabNumber = 2
        inputFieldPath = 2
        tableHeaderPath = '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[4]/transform/div/div[3]/div/div/visual-modern/div/div/div[2]/div[1]/div[1]/div/div/div'
        totalNoOfColumns = 13
        tableHeaderWithData = ["TYPE", "BRANCH", "ENTITYCODE", "INVOICE NO", "DOC_DT", "SLMASTER", "FINALAMOUNT", "DESCRIPTION", "TKT NO", "PAX NAME", "CLIENT NAME", "OTHER REMARKS", "BRANCH_1"]

    elif tab == 'bsp-statement':
        tabNumber = 4
        inputFieldPath = 1
        tableHeaderPath = '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[6]/transform/div/div[3]/div/div/visual-modern/div/div/div[2]/div[1]/div[1]/div/div/div'
        totalNoOfColumns = 18
        tableHeaderWithData = ["TKTT TYPE", "Ticket No", "Sum of Gross Amount", "BSP FOP", "Airline Name", "Agent (incl Check Digit)", "Agent IATA Region", "Type Group", "RA NO", "Date of Issue", "Billing Period", "Credit Card Number (masked)", "Passenger Name", "Passenger Last Name", "Total Commission", "Penalty", "Balance Payable", "PNR"]
    
    elif tab == 'bsp-travcom':
        tabNumber = 5
        inputFieldPath = 2
        tableHeaderPath = '//*[@id="pvExplorationHost"]/div/div/exploration/div/explore-canvas/div/div[2]/div/div[2]/div[2]/visual-container-repeat/visual-container[4]/transform/div/div[3]/div/div/visual-modern/div/div/div[2]/div[1]/div[1]/div/div/div'
        totalNoOfColumns = 19
        tableHeaderWithData = ["TKTT TYPE", "Ticket No", "Gross Amount", "Payable", "Total Comm", "FOP", "EMP ID", "PNR NO", "InvoiceNumber", "MainTag", "InvoiceDate", "Division", "Div_Name", "ProfileName", "ProfileName", "ValidatingCarrier", "TicketingAgentName", "IataNumber", "IataName"]
    
    entireTableData = automation(driver, tabNumber, inputFieldPath, tableHeaderPath, totalNoOfColumns, tableHeaderWithData)
    return entireTableData

# Header for excel file
data = findTab(tab)

# Create an Excel workbook and sheet
workbook = openpyxl.Workbook()
sheet = workbook.active
sheet.title = "Extracted Data"

# Using limit, setting columns and adding the extracted data to sheet
for i in range(0,len(data), totalNoOfColumns):
    sheet.append(data[i: i+totalNoOfColumns])

# Specify the directory where you want to save the file
directory = 'C:\\Users\\ganesh.ss\\Downloads'

# Construct the full file path
file_path = os.path.join(directory, fileName)

# Increment the counter if the file already exists
base_filename = file_path

# Checking for extension
if not base_filename.endswith('.xlsx'):
    base_filename += '.xlsx'

# Writing checking filename and renaming filename
counter = 0
while os.path.exists(base_filename):
    counter += 1
    if not fileName.endswith('.xlsx'):
        fileName += '.xlsx'
    temp = fileName.split('.')
    base_filename = f"{temp[0]}_{counter}.{temp[1]}"
    base_filename = os.path.join(directory, base_filename)

# Save the Excel file
workbook.save(base_filename)
print(base_filename + " created successfully.")